import pandas as pd
from openpyxl import load_workbook

"""
import of pandas to work with data and load_workbook to work with MS_Excel
"""

fn = r'C:\Users\User\Downloads\2022_new.xlsx'
wb = load_workbook(fn)
ws = wb['победители']
df_original = pd.read_excel(fn, sheet_name='номинанты')
nominees = pd.read_excel(fn, sheet_name='списки')

top = [10, 9, 8, 7, 6, 5, 4, 3, 2, 1]
nomination = [
    "director",
    "actor",
    "actress",
    "actor2",
    "actress2",
    "original_screenplay",
    "adapted_screenplay",
    "operator",
    "editing",
    "soundtrack",
    "song",
    "art_direction",
    "costumes",
    "make_up",
    "effects",
    "sound",
    "stunts",
    "animation",
    "documentation",
    "russian",
    "live_action_short",
    "animated_short",
    "documentary_short",
    "debut",
    "ensemble",
    "using_music",
    "young_actor",
    'special_mentions',
    "choreography"
]

nomination_plus = ["movie"] + nomination

df_movies = pd.DataFrame()
df_first = pd.DataFrame()
df_second = pd.DataFrame()
df_third = pd.DataFrame()
data = pd.DataFrame()
count = 0


def delete_non_breaking_spaces(text):
    return text.replace(u'\xa0', u' ').strip()


def results(data):
    point = []
    for x in range(len(data)):
        point.append(data[x])
    return point


def actors_sovpad(data,data2,actor,actor2,points,points2):
    data_sovpad=pd.DataFrame()
    sovpad_counter=0
    for a_index in range(len(df_second)):
        for a2_index in range(len(df_second)):
            if data.loc[a_index,actor]==data2.loc[a2_index,actor2]:
                sovpad_counter+=1
                data_sovpad.loc[sovpad_counter,'actor'] = data.loc[a_index,actor]
                data_sovpad.loc[sovpad_counter,'points'] = data.loc[a_index,points]+data2.loc[a2_index,points2]
    return data_sovpad


df_original = df_original.drop(columns = 'Отметка времени').set_index('Ваш ник на Форуме Кинопоиска:')
df_first = df_original.T
nominees.columns = nomination

for name, values in df_first .items():
    """
    create new table with movies only, order matters
    """
    values = values.dropna().reset_index(drop=True)

    for x in range(0,10):
        data.loc[x, name] = values[x]

for name, values in data.items():
    """
    create new table with  users, movies, and points
    """
    for x in range(len(top)):
        df_movies.loc[count, "user"] = name
        df_movies.loc[count, "movie"] = data.loc[x, name]
        df_movies.loc[count, "point"] = top[x]
        count += 1

for nom in range(len(nomination)):
    count_nomination = 0
    nominee = nominees.loc[:, nomination[nom]].dropna().reset_index(drop=True)
    for a in range(len(nominee)):
        clean_nominee = delete_non_breaking_spaces(nominee[a])
        df_second.loc[count_nomination, "best_{}".format(nomination[nom])] = clean_nominee
        points = 0
        names = []
        for name, values in df_first.items():
            values = values.fillna('xxx').reset_index(drop=True)
            if clean_nominee in values[10+nom]:
                points +=1
                names.append(name)
        df_second.loc[count_nomination, 'points_{}'.format(nomination[nom])] = points
        df_second.loc[count_nomination, 'users_{}'.format(nomination[nom])] = ', '.join(names)
        count_nomination += 1

df_best = df_movies.pivot_table(
    index="movie", values="point", aggfunc="sum"
).reset_index(level="movie")
best_movies = (
    df_best.groupby("movie")["point"].sum().sort_values(ascending=False).reset_index()
)

for movie_number in range(len(best_movies['movie'])):
    mention = []
    for x in range(len(df_movies['movie'])):
        if best_movies.loc[movie_number, 'movie'] in df_movies.loc[x,'movie']:
            if df_movies.loc[x,'point'] == 1:
                point = 'балл'
            elif df_movies.loc[x,'point'] > 4:
                point = 'баллов'
            else:
                point = 'балла'
            mention.append('{} - {} {}'.format(df_movies.loc[x,'user'], df_movies.loc[x,'point'].astype(int), point))
        best_movies.loc[movie_number, 'mentions'] = ', '.join(mention)

for x in range(len(nomination)):
    df_third = pd.DataFrame()
    df_third["{}".format(nomination[x])] = df_second.pivot_table(
        index="best_{}".format(nomination[x]),
        values="points_{}".format(nomination[x]),
        aggfunc="sum",
    )
    df_third = df_third.reset_index(level="best_{}".format(nomination[x]))
    df_second = df_second.sort_values(by =["best_{}".format(nomination[x])])
    df_second = df_second.reset_index(drop=True)
    df_third['users_{}'.format(nomination[x])] = df_second['users_{}'.format(nomination[x])]
    a = (
        df_third[["best_{}".format(nomination[x]), "{}".format(nomination[x]), 'users_{}'.format(nomination[x])]]
        .sort_values(by="{}".format(nomination[x]), ascending=False)
        .reset_index()
    )
    value = results(a["best_{}".format(nomination[x])])
    value_point = results(a["{}".format(nomination[x])])
    value_mentions = results(a['users_{}'.format(nomination[x])])
    for y in range(len(value)):
        cell = ws.cell(row=y + 2, column=4 + (x * 3))
        cell_point = ws.cell(row=y + 2, column=5 + (x * 3))
        cell_mentions = ws.cell(row=y + 2, column=6 + (x * 3))
        cell.value = value[y]
        cell_point.value = value_point[y]
        cell_mentions.value = value_mentions[y]

dfs_to_write = []
startrow = 0
writer = pd.ExcelWriter(fn)
main=["actor","actress"]
support=["actor2","actress2"]

for index in range(len(main)):
    actor = 'best_{}'.format(main[index])
    actor2 = 'best_{}'.format(support[index])
    points = 'points_{}'.format(main[index])
    points2 = 'points_{}'.format(support[index])
    df_main = df_second[['best_{}'.format(main[index]),'points_{}'.format(main[index])]]
    df_support = df_second[['best_{}'.format(support[index]),'points_{}'.format(support[index])]]
    data_coincidences = actors_sovpad(df_main, df_support,actor,actor2,points,points2)
    print(data_coincidences)

for df in dfs_to_write:
    df.to_excel(writer, sheet_name='совпадения', startrow=startrow, index=False)
    startrow += len(df) + 2

for x in range(len(best_movies["movie"])):
    value = best_movies.loc[x, "movie"]
    value_point = best_movies.loc[x, "point"]
    value_mentions = best_movies.loc[x, 'mentions']
    cell = ws.cell(row=x + 2, column=1)
    cell_point = ws.cell(row=x + 2, column=2)
    cell_mentions = ws.cell(row=x + 2, column=3)
    cell.value = value
    cell_point.value = value_point
    cell_mentions.value = value_mentions

for z in range(len(nomination_plus)):
    cell = ws.cell(row=1, column=1 + (z * 3))
    cell_point = ws.cell(row=1, column=2 + (z * 3))
    cell_mentions = ws.cell(row=1, column=3 + (z * 3))
    cell.value = nomination_plus[z]
    cell_point.value = "points"
    cell_mentions.value = 'mentions_by'

wb.save(fn)
wb.close()