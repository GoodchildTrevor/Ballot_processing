import pandas as pd
from openpyxl import load_workbook

fn = r'C:\Users\user\Downloads\1999.xlsx'
wb = load_workbook(fn)
ws = wb['Sheet1']
df_original = pd.read_excel(r'C:\Users\user\Downloads\1999.xlsx', sheet_name='номинанты')

top = [10, 9, 8, 7, 6, 5, 4, 3, 2, 1]
beginning0 = ["10.", "1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9."]
nomination = [
    "director",
    "actor",
    "actress",
    "actor2",
    "actress2",
    "original_screenplay",
    "adapted_screenplay",
]
nomination_plus = ["movie"] + nomination

df_movies = pd.DataFrame()
df_second = pd.DataFrame()
df_third = pd.DataFrame()
data = pd.DataFrame()
count = 0


def change(value1, frame, attr):
    """
    Merge entities if one is part of another.
    :param value1: first value
    :param frame: frame we are searching in
    :param attr: attribute we are looking for
    :return: final s
    """
    # TODO: complexity is too high
    # TODO: if values are equal, iteration will end
    for x in range(len(frame[attr])):
        value2 = frame.loc[x, attr]
        if value1 in value2:
            return value2
        elif value2 in value1:
            return value1


def replacer(a):
    a = a.replace("(", "")
    a = a.replace(")", "")
    a = a.replace("ё", "е")
    a = a.replace(",", "")
    a = a.replace("  ", " ")
    a = a.replace("-", "")
    for index in range(len(top)):
        a = a.replace(beginning0[index], "")
    return a


def results(data):
    point = []
    for x in range(len(data)):
        point.append(data[x])
    return point


def prob(movie):
    movie = movie.strip()
    return movie


for name, values in df_original.items():
    values = values.dropna().reset_index(drop=True)
    for x in range(0, 11):
        data.loc[x, name] = values[x]

for y in range(len(nomination)):
    count_nomination = 0
    for name, values in df_original.items():
        values = values.dropna().reset_index(drop=True)
        for z in range(12 + (y * 6), 17 + (y * 6)):
            df_second.loc[count_nomination, "user_{}".format(nomination[y])] = name
            df_second.loc[count_nomination, "best_{}".format(nomination[y])] = values[z]
            df_second.loc[count_nomination, "point_{}".format(nomination[y])] = 1
            count_nomination += 1

for name, values in data.items():
    for x in range(len(top)):
        df_movies.loc[count, "user"] = name
        df_movies.loc[count, "movie"] = data.loc[x + 1, name]
        df_movies.loc[count, "point"] = top[x]
        count += 1

for x in range(len(nomination)):
    df_second["best_{}".format(nomination[x])] = df_second[
        "best_{}".format(nomination[x])
    ].str.lower()
    df_second["best_{}".format(nomination[x])] = df_second[
        "best_{}".format(nomination[x])
    ].apply(replacer)
    df_second["best_{}".format(nomination[x])] = df_second[
        "best_{}".format(nomination[x])
    ].apply(prob)
    df_second['best_{}'.format(nomination[x])] = df_second[
        'best_{}'.format(nomination[x])].apply(
        change, args=(df_second, 'best_{}'.format(nomination[x])))

df_movies["movie"] = df_movies["movie"].str.lower()
df_movies["movie"] = df_movies["movie"].apply(replacer)
df_movies["movie"] = df_movies["movie"].apply(prob)
df_movies["movie"] = df_movies["movie"].apply(change, args=(df_movies, "movie"))

df_second["best_director"] = df_second["best_director"].apply(
    change, args=(df_second, "best_director")
)
df_second["best_actor"] = df_second["best_actor"].apply(change_actor)
df_second["best_actress"] = df_second["best_actress"].apply(change_actress)
df_second["best_actor2"] = df_second["best_actor2"].apply(change_actor2)
df_second["best_actress2"] = df_second["best_actress2"].apply(change_actress2)
df_second["best_original_screenplay"] = df_second["best_original_screenplay"].apply(
    change_original_screenplay
)
df_second["best_adapted_screenplay"] = df_second["best_adapted_screenplay"].apply(
    change_adapted_screenplay
)

for x in range(len(nomination)):
    df_third = pd.DataFrame()
    df_third["{}".format(nomination[x])] = df_second.pivot_table(
        index="best_{}".format(nomination[x]),
        values="point_{}".format(nomination[x]),
        aggfunc="count",
    )
    df_third = df_third.reset_index(level="best_{}".format(nomination[x]))
    df_third = df_third[df_third["best_{}".format(nomination[x])] != "zzz"]
    a = (
        df_third[["best_{}".format(nomination[x]), "{}".format(nomination[x])]]
        .sort_values(by="{}".format(nomination[x]), ascending=False)
        .reset_index()
    )
    value = results(a["best_{}".format(nomination[x])])
    value_point = results(a["{}".format(nomination[x])])
    for y in range(len(value)):
        cell = ws.cell(row=y + 2, column=3 + (x * 2))
        cell_point = ws.cell(row=y + 2, column=4 + (x * 2))
        cell.value = value[y]
        cell_point.value = value_point[y]

df_best = df_movies.pivot_table(
    index="movie", values="point", aggfunc="sum"
).reset_index(level="movie")
best_movies = (
    df_best.groupby("movie")["point"].sum().sort_values(ascending=False).reset_index()
)

for x in range(len(best_movies["movie"])):
    value = best_movies.loc[x, "movie"]
    value_point = best_movies.loc[x, "point"]
    cell = ws.cell(row=x + 2, column=1)
    cell_point = ws.cell(row=x + 2, column=2)
    cell.value = value
    cell_point.value = value_point

for z in range(len(nomination_plus)):
    cell = ws.cell(row=1, column=1 + z * 2)
    cell_point = ws.cell(row=1, column=2 + z * 2)
    cell.value = nomination_plus[z]
    cell_point.value = "points"

wb.save(fn)
wb.close()
